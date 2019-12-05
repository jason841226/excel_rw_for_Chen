import openpyxl
first_idx = 30316
n_data = 3
output_filename = argv[1]
chart_title = argv[2]
colors = ['green','grey','deepSkyBlue','aqua','dkCyan','red']
#available_color = ['lavender', 'peachPuff', 'plum', 'firebrick', 'mediumAquamarine', 'dkGoldenrod', 'mediumOrchid', 'powderBlue', 'orangeRed', 'darkBlue', 'skyBlue', 'green', 'greenYellow', 'ltGray', 'seaShell', 'violet', 'dimGray', 'chartreuse', 'dkRed', 'lightGoldenrodYellow', 'ltPink', 'medBlue', 'peru', 'teal', 'mintCream', 'dkSlateGray', 'navajoWhite', 'maroon', 'mediumSlateBlue', 'darkGrey', 'darkSalmon', 'dkOliveGreen', 'dodgerBlue', 'steelBlue', 'cadetBlue', 'paleGoldenrod', 'linen', 'mediumBlue', 'darkSlateBlue', 'silver', 'mediumVioletRed', 'blueViolet', 'cornsilk', 'hotPink', 'lightSalmon', 'blanchedAlmond', 'purple', 'yellow', 'lightSlateGray', 'lawnGreen', 'indigo', 'darkCyan', 'slateBlue', 'ltCyan', 'mediumPurple', 'paleGreen', 'darkSeaGreen', 'dkSeaGreen', 'chocolate', 'ltBlue', 'rosyBrown', 'sandyBrown', 'dkOrchid', 'dkBlue', 'red', 'beige', 'ltSlateGray', 'ltSeaGreen', 'dkOrange', 'cyan', 'midnightBlue', 'dimGrey', 'medSeaGreen', 'dkViolet', 'medAquamarine', 'lightGreen', 'darkViolet', 'dkMagenta', 'darkKhaki', 'dkKhaki', 'limeGreen', 'white', 'yellowGreen', 'medTurquoise', 'forestGreen', 'darkGreen', 'ghostWhite', 'burlyWood', 'medSlateBlue', 'cornflowerBlue', 'floralWhite', 'indianRed', 'ltGrey', 'lightPink', 'oldLace', 'medVioletRed', 'dkGrey', 'ltSalmon', 'mediumSeaGreen', 'bisque', 'medPurple', 'darkOliveGreen', 'antiqueWhite', 'medOrchid', 'olive', 'honeydew', 'lightSkyBlue', 'mediumTurquoise', 'lightYellow', 'aliceBlue', 'lavenderBlush', 'oliveDrab', 'medSpringGreen', 'slateGrey', 'crimson', 'darkTurquoise', 'dkTurquoise', 'paleTurquoise', 'orange', 'saddleBrown', 'darkOrange', 'azure', 'khaki', 'lime', 'papayaWhip', 'dkSalmon', 'mediumSpringGreen', 'lightGray', 'ltGreen', 'darkOrchid', 'lightSlateGrey', 'darkGoldenrod', 'dkCyan', 'lightBlue', 'magenta', 'gold', 'royalBlue', 'fuchsia', 'ltYellow', 'darkSlateGrey', 'darkMagenta', 'snow', 'coral', 'ltSkyBlue', 'lightCoral', 'dkSlateBlue', 'mistyRose', 'ltSlateGrey', 'slateGray', 'paleVioletRed', 'orchid', 'black', 'brown', 'dkGreen', 'dkGray', 'lightCyan', 'ltGoldenrodYellow', 'seaGreen', 'ivory', 'lightSteelBlue', 'blue', 'darkGray', 'aquamarine', 'darkSlateGray', 'ltSteelBlue', 'darkRed', 'sienna', 'thistle', 'dkSlateGrey', 'gainsboro', 'tomato', 'pink', 'turquoise', 'deepPink', 'gray', 'lightSeaGreen', 'tan', 'ltCoral', 'springGreen', 'navy', 'lightGrey', 'deepSkyBlue', 'lemonChiffon', 'aqua', 'grey', 'goldenrod', 'whiteSmoke', 'wheat', 'salmon', 'moccasin']

characters = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']

#read template
wb = openpyxl.load_workbook('template.xlsx')
ws = wb[str(wb.sheetnames[0])]

for i in range(n_data):
	#create a new worksheet with a copy of the template
	ws2 = ws if i==0 else wb.copy_worksheet(ws)
	ws2.title = str(first_idx+i)

	#read data from 30316.xlsx...
	wb_data = openpyxl.load_workbook(str(first_idx+i) + '.xlsx')
	ws_data = wb_data[str(wb_data.sheetnames[0])]

	#write data to the new worksheet
	for row in ws_data.rows:
		for cell in row:
			idx = characters[cell.column-1]+str(cell.row)
			if cell.value != None:
				ws2[idx] = cell.value

	#draw line chart with data
	c1 = openpyxl.chart.LineChart()
	c1.title = chart_title
	for j in range(6):
		data = openpyxl.chart.Reference(ws2, min_col=26+j, min_row=9, max_row=609)
		s = openpyxl.chart.Series(data, title_from_data=True)
		s.graphicalProperties.line = openpyxl.drawing.line.LineProperties(solidFill = openpyxl.drawing.colors.ColorChoice(prstClr=colors[j]))
		c1.series.append(s)
	x_data = openpyxl.chart.Reference(ws2,min_col=1,min_row=2,max_row=601)
	c1.set_categories(x_data)

	ws2.add_chart(c1,"AG2")

wb.save(output_filename)