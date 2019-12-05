import openpyxl
input_filename = argv[1]
output_filename = argv[2]

cmp_list = ['Z2','AA2','AB2','AC2','AD2','AE2']
characters = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']

group_number = [2,1]

max_values=[]
wb = openpyxl.load_workbook(input_filename,data_only=True)
for ws_name in wb.sheetnames:
	ws = wb[str(ws_name)]
	max_values.append(max([ws[idx].internal_value for idx in cmp_list]))

print(max_values)

wb_write = openpyxl.Workbook()
ws_write = wb_write.active
abc=0
idx2=0
for n in group_number:
	for i in range(n):
		idx = characters[abc]+str(i+1)
		ws_write[idx] = max_values[idx2]
		idx2 = idx2 + 1
	abc = abc + 1
wb_write.save(output_filename)